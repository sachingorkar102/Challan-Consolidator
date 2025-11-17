
import pdfplumber
import pandas as pd
from dataclasses import asdict
import re
import os
from pathlib import Path
from challan_info import ChallanInfo
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime


DEFAULT_SECTION_MAP = {"94C": "194C - Works Contract","94H": "194H - Commission / Brokerage",
               "94I":"194I(b) - Land / Building rent",
               "94J":"194J - Fees / Royalty (Others)",
               "94Q": "194Q - Purchase of goods"}

def get_int(string: str) -> int:
    return int(re.sub(r"[^\d]", "", string))

def get_section(string: str,section_map: dict) -> str:
    if string in section_map:
        return section_map[string]
    else:
        return string

def remove_duplicate_challans(challans: ChallanInfo):
    unique_challans = []
    seen = set()
    duplicates = set()

    for c in challans:
        if c.challan_no in seen:
            print(f"Duplicate found: File Name {c.file_name}, removing challan from list...")
            duplicates.add(c.challan_no)
        else:
            unique_challans.append(c)
            seen.add(c.challan_no)

    return unique_challans

def format_excel(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active

    # Format header row
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, horizontal="center")

    # Auto-fit column widths based on max content length
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter  # e.g. 'A', 'B'
        for cell in column_cells:
            try:
                cell_value = str(cell.value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(excel_file)

def load_map(file_path: str, default_map: dict):
    path = file_path/"section_map.txt"
    if not path.exists():
        return default_map
    result = {}
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or "=" not in line:
                continue 
            key, value = line.split("=", 1) 
            result[key.strip()] = value.strip()

    return result

def format_indian(num):
    num_str = str(num)
    # Handle decimal part
    if '.' in num_str:
        whole, dec = num_str.split('.')
    else:
        whole, dec = num_str, None

    # First group (last 3 digits)
    last3 = whole[-3:]
    rest = whole[:-3]

    # Add commas every 2 digits
    if rest:
        rest = list(rest)
        groups = []
        while len(rest) > 2:
            groups.insert(0, ''.join(rest[-2:]))
            rest = rest[:-2]
        if rest:
            groups.insert(0, ''.join(rest))
        formatted = ",".join(groups) + "," + last3
    else:
        formatted = last3

    if dec:
        formatted += "." + dec

    return formatted

    

folder_path = Path.cwd()
pdf_files = [f for f in folder_path.glob("*.pdf")]
challans = []
section_map = load_map(folder_path,DEFAULT_SECTION_MAP)
count = 0
for pdf in pdf_files:
    try:
        count += 1
        data = {}
        
        with pdfplumber.open(pdf) as ppdf:
            page = ppdf.pages[0]
            tables = page.extract_tables()
            # FOR TABLE 1
            df1 = pd.DataFrame(tables[0],columns=["line"])
            data_map = {}
            for index, row in df1.iterrows():
                if ":" in row["line"]:
                    key, value = row["line"].split(":", 1)
                    data_map[key.strip()] = value.strip()
            
            
            data["file_name"] = pdf.name
            data["challan_no"] = data_map.get("Challan No")
            data["bsr_code"] = data_map.get("BSR code")
            data["tender_date"] = data_map.get("Tender Date")
            
            section_raw = data_map.get("Nature of Payment", "")
            data["section_raw"] = section_raw
            data["section"] = get_section(section_raw,section_map)
            data["financial_year"] = data_map.get("Financial Year")
            data["total_amount"] = get_int(data_map.get("Amount (in Rs.)"))
            data["firm_name"] = data_map.get("Name")


            # FOR TABLE 2
            df2 = pd.DataFrame(tables[1],columns=["line"])
            tax_amount = 0
            interest_amount = 0
            other_amount = 0
            for _, row in df2.iterrows():   # table2_df is your Table 2 DataFrame
                line = row["line"].strip()

                if line.startswith("A Tax"):
                    tax_amount = get_int(line.replace("A Tax",""))

                elif line.startswith("D Interest"):
                    interest_amount = get_int(line.replace("D Interest",""))

                elif line.startswith("B Surcharge"):
                    other_amount += get_int(line.replace("B Surcharge",""))

                elif line.startswith("C Cess"):
                    other_amount += get_int(line.replace("C Cess",""))

                elif line.startswith("E Penalty"):
                    other_amount += get_int(line.replace("E Penalty",""))

                elif line.startswith("F Fee under section 234E"):
                    other_amount += get_int(line.replace("F Fee under section 234E",""))
            data["tax_amount"] = tax_amount
            data["interest_amount"] = interest_amount
            data["other_amount"] = other_amount
            if(tax_amount+interest_amount+other_amount != data["total_amount"]):
                print("Error Exctracting data from challan file "+pdf.name)
                continue
            challaninfo = ChallanInfo(**data)
            challans.append(challaninfo)
            print(f"\rData extracted from {count} out of {len(pdf_files)} Challan files.", end="")
    except Exception as e:
        print(f" and error occured while procedding pdf {pdf.name}")
        print(e)

print("")
challans = remove_duplicate_challans(challans)
if(len(challans) != 0):
    records = [asdict(c) for c in challans]
    df = pd.DataFrame(records)
    desired_columns = ["tender_date","section","tax_amount","interest_amount","fee_amount","other_amount","total_amount","book_entry","challan_no","bsr_code","sa_entry","financial_year","firm_name"]
    df = df[desired_columns]
    df.rename(columns={
    "tender_date": "Date of Challan",
    "financial_year": "Financial Year",
    "challan_no": "Challan Serial No.",
    "bsr_code": "Bank Branch Code",
    "section": "Section",
    "total_amount": "Total Amount Deposited",
    "tax_amount": "Deposited - Tax",
    "interest_amount": "Interest",
    "fee_amount": "Fee",
    "other_amount": "Other Amount",
    "book_entry": "Book - Entry ?",
    "sa_entry": "Type of Payment",
    "firm_name": "Name"
    }, inplace=True)
    
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    excel_file = f"challans_{timestamp}.xlsx"
    print("Saving details to excel...")
    df.to_excel(excel_file, index=False)
    format_excel(excel_file)
    file_count = 0
    for challan in challans:
        try:
            file_count += 1
            os.rename(challan.file_name,challan.section_raw+" "+str(format_indian(challan.total_amount))+" "+challan.challan_no+".pdf")
            print(f"\rFiles renamed {file_count} out of {len(challans)} Challan files.", end="")
        except Exception as e:
            print(f"Error renaming {challan.file_name}: {e}")
else:
    print("No Challans found in folder...")
print("")
input("Press Enter to exit...")